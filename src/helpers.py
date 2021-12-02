import os, glob
from pathlib import Path

import pandas as pd
import matplotlib.pyplot as plt
import xlrd, xlwt, xlutils
import openpyxl

from config import STATIC_DIR

def get_header(sheet):
    col_names = []
    for column in sheet.iter_cols(1, sheet.max_column):
        col_names.append(column[0].value)
    return col_names

def print_some_rows(sheet):
    for row in sheet.iter_rows(max_row = 3):
        for cell in row:
            print(cell.value, end = ' ')
        print()

get_docs = lambda: glob.glob(os.path.join(STATIC_DIR, '*.xlsx'))

def main():
    available_docs = get_docs()
    if len(available_docs) > 0:
        xlsx_file = Path(available_docs[0])
        wb_obj = openpyxl.load_workbook(xlsx_file)

        sheet = wb_obj.active
        print(f'Value of C2 cell: {sheet["C2"].value}')
        print('Header:', get_header(sheet))
        print_some_rows(sheet)
        print(f'Rows: {sheet.max_row}, Columns: {sheet.max_column}')
        sheet, wb_obj = None, None

        df = pd.read_excel(available_docs[0], index_col = 0)
        print(df.head())
        print(df.shape)

        xlsx = pd.ExcelFile(available_docs[0])
        sheets = pd.concat([xlsx.parse(sheet) for sheet in xlsx.sheet_names])
        print(df.tail())
        res = df.sort_values(['Country'], ascending = False).head(5)
        print(res)
        print(res.describe())
        print(res['Year'].mean())
        print(df[['Year', 'Country']])
        # res.plot(kind = 'bar', figsize = (20, 8))
        # plt.show()
        # df.to_excel('output.xlsx')

if __name__ == '__main__':
    main()